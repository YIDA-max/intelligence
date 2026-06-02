"""
菜单导入 xlsx 生成脚本
用法：python doc/gen_menu.py

根据输入的菜单路径和权限配置，基于模板格式生成可直接导入的 xlsx 文件。
"""

import shutil
import os
import openpyxl
from openpyxl import load_workbook
from copy import copy

TEMPLATE_PATH = "docAI/单个菜单模板 - 副本.xlsx"

# ─────────────────────────────────────────────
# 配置区：每次生成新模块只需修改这里
# ─────────────────────────────────────────────
CONFIG = {
    "menu_path": "财务-来往记录-请款池",
    "route": "/financial/transactionManagement/paymentRequestPool/index",
    "menu_sort": 10,
    "icon": "iconfont icon-tenant",
    "mid_routes": {
        "来往记录": "/financial/transactionManagement/index",
    },
    "module_key": "paymentRequestPool",
    "buttons": [
        ("查看", "view", 10),
        ("导出", "export", 20),
        ("作废", "discard", 30),
        ("日志", "log", 40),
        ("请款", "request", 50),
    ],
    "output": "docAI/output_menu.xlsx",
}
# ─────────────────────────────────────────────


def copy_cell_style(src_cell, dst_cell):
    """复制单元格样式"""
    if src_cell.has_style:
        dst_cell.font      = copy(src_cell.font)
        dst_cell.fill      = copy(src_cell.fill)
        dst_cell.border    = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


def get_style_row(ws, row_idx):
    """获取某行的样式（用于复制到新行）"""
    return [ws.cell(row=row_idx, column=c) for c in range(1, 14)]


def build_rows(cfg):
    """
    根据配置构建数据行列表
    每行格式：(ID, 类型, 上级菜单, 名称, 排序, 权限标识, 路由, 图标, 缓冲, 显示, 内嵌, 带参, 组件)
    """
    parts = cfg["menu_path"].split("-")
    # 至少需要 一级-二级-三级，从三级开始生成（二级默认已存在）
    if len(parts) < 3:
        raise ValueError("menu_path 至少需要三级，如：仓库-库存-WFS进销存报表")

    # 从最终路由中推导各中间层级的路由前缀
    # 例如 route=/warehouse/stockManage/wfsInventoryReport/index，共3段（去掉/index后）
    # 中间层级数 = len(parts) - 2（跳过首尾）
    route_without_index = cfg["route"].rsplit("/index", 1)[0]  # /warehouse/stockManage/wfsInventoryReport
    route_segments = route_without_index.strip("/").split("/")  # ['warehouse', 'stockManage', 'wfsInventoryReport']

    # 支持手动指定中间层级路由，格式：{ "菜单名": "/path/index" }
    mid_routes_override = cfg.get("mid_routes", {})

    rows = []

    # 从第三级开始生成（跳过第一级和第二级，因为二级默认已存在）
    # 如果用户明确要求生成二级菜单，可以在 CONFIG 中设置 "skip_second_level": False
    skip_second_level = cfg.get("skip_second_level", True)
    start_index = 2 if skip_second_level else 1

    for i in range(start_index, len(parts)):
        name   = parts[i]
        parent = parts[i - 1]
        is_last = (i == len(parts) - 1)

        if is_last:
            # 最后一级：菜单类型，带完整路由
            rows.append((
                None, "菜单", parent, name,
                cfg["menu_sort"], None,
                cfg["route"], cfg["icon"],
                "否", "是", "否", "否", None
            ))
        else:
            # 中间层级：优先使用 mid_routes 手动指定，否则从路由段推导
            if name in mid_routes_override:
                mid_route = mid_routes_override[name]
            else:
                seg_index = i - 1
                if seg_index < len(route_segments):
                    mid_route = "/" + "/".join(route_segments[:seg_index + 1]) + "/index"
                else:
                    mid_route = None
            rows.append((
                None, "菜单", parent, name,
                10, None,
                mid_route, cfg["icon"],
                "否", "是", "否", "否", None
            ))

    # 按钮权限行（上级菜单 = 最后一级菜单名）
    last_menu = parts[-1]

    # 支持 button_groups（多moduleKey场景）和 buttons（单moduleKey场景）
    button_groups = cfg.get("button_groups") or []
    if not button_groups and cfg.get("buttons"):
        # 兼容旧格式：单 module_key + buttons
        button_groups = [{"module_key": cfg["module_key"], "buttons": cfg["buttons"]}]

    for group in button_groups:
        mk = group["module_key"]
        for btn_name, btn_suffix, btn_sort in group["buttons"]:
            rows.append((
                None, "按钮", last_menu, btn_name,
                btn_sort, f"{mk}_{btn_suffix}",
                None, None,
                "否", "是", "否", "否", None
            ))

    return rows


def generate(cfg):
    # 复制模板
    shutil.copy(TEMPLATE_PATH, cfg["output"])
    wb = load_workbook(cfg["output"])
    ws = wb.active

    # 读取模板数据行的样式（第3行，即 row index 3）
    template_data_row = get_style_row(ws, 3)

    # 清除模板中已有的数据行（保留第1行空行 + 第2行表头）
    # 模板数据从第3行开始
    max_row = ws.max_row
    for r in range(3, max_row + 1):
        for c in range(1, 14):
            ws.cell(row=r, column=c).value = None

    # 写入新数据行
    data_rows = build_rows(cfg)
    for i, row_data in enumerate(data_rows):
        row_idx = 3 + i
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            copy_cell_style(template_data_row[col_idx - 1], cell)
        # 设置行高
        ws.row_dimensions[row_idx].height = 15.0

    wb.save(cfg["output"])
    print(f"✓ 生成成功：{cfg['output']}")
    print(f"  共 {len(data_rows)} 行数据：")
    for r in data_rows:
        print(f"    [{r[1]}] {r[3]}  上级:{r[2]}  权限:{r[5] or '-'}  路由:{r[6] or '-'}")


if __name__ == "__main__":
    generate(CONFIG)
